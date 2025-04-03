import os
import sys
import time
import statistics
from openpyxl import load_workbook, Workbook

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))

from shared.data_loader.data_loader import load_data
from shared.utils import evaluate_solution, verify_solution

from random_method.heuristics import evolutionary_one_plus_one
from constructive_method.heuristics import (
    nearest_neighbor_minimize_max_workload_time,
    nearest_neighbor_minimize_max_workload_time_randomized
)

# Parameters
instances_list = [
    '40_homogeneous.xlsx', 
    '40_heterogeneous.xlsx', 
    '60_homogeneous.xlsx',
    '60_heterogeneous.xlsx',
    '80_homogeneous.xlsx',
    '80_heterogeneous.xlsx',
]

n_random_runs = 30
n_random_iterations = 1000
n_evolutionary_runs = 30
evolutionary_max_iterations = 1000

bks_file = 'analysis/find_bks/bks_results.xlsx'

def load_bks():
    wb = load_workbook(bks_file)
    ws = wb.active
    return {
        row[0].value: {
            'bks_wmax': float(row[1].value),
            'bks_wmax_wmin': float(row[2].value)
        }
        for row in ws.iter_rows(min_row=2)
    }

def evaluate_method(method_name, func, args, runs, method_kwargs=None):
    wmax_values = []
    wmax_wmin_values = []
    execution_times = []

    for _ in range(runs):
        start = time.time()
        result = func(*args, **(method_kwargs or {}))
        assignments, load_zones, _ = result
        elapsed_time = time.time() - start

        wmax, wmax_wmin = evaluate_solution(load_zones)

        if verify_solution(assignments, load_zones, *args[:3]):
            wmax_values.append(wmax)
            wmax_wmin_values.append(wmax_wmin)
            execution_times.append(elapsed_time)

    return {
        'method': method_name,
        'best_wmax': round(min(wmax_values), 2),
        'mean_wmax': round(statistics.mean(wmax_values), 2),
        'best_wmax_wmin': round(min(wmax_wmin_values), 2),
        'mean_wmax_wmin': round(statistics.mean(wmax_wmin_values), 2),
        'time_sec': round(statistics.mean(execution_times), 4)
    }

def compute_gap(mean, bks):
    return round(((mean - bks) / bks) * 100, 2) if bks != 0 else 0.0

def main():
    bks_dict = load_bks()

    wb = Workbook()
    ws = wb.active
    ws.title = 'comparison'
    ws.append([
        'instance', 'method',
        'bks_wmax', 'bks_wmax_wmin',
        'best_wmax', 'mean_wmax', 'gap_wmax_percent',
        'best_wmax_wmin', 'mean_wmax_wmin', 'gap_wmax_wmin_percent',
        'time_sec'
    ])

    for instance in instances_list:
        print(f"Processing instance: {instance}")
        args = load_data(instance)
        bks = bks_dict.get(instance)
        if not bks:
            print(f"⚠️ BKS not found for {instance}, skipping.")
            continue

        # Deterministic method
        det_result = evaluate_method(
            'deterministic',
            nearest_neighbor_minimize_max_workload_time,
            args, 1
        )

        # Randomized method
        rand_result = evaluate_method(
            'randomized',
            nearest_neighbor_minimize_max_workload_time_randomized,
            args, n_random_runs,
            {'N': n_random_iterations}
        )

        # Evolutionary 1+1 method
        evo_result = evaluate_method(
            'evolutionary_1_plus_1',
            evolutionary_one_plus_one,
            args, n_evolutionary_runs,
            {'max_iterations': evolutionary_max_iterations}
        )

        for res in [det_result, rand_result, evo_result]:
            gap_wmax = compute_gap(res['mean_wmax'], bks['bks_wmax'])
            gap_wmax_wmin = compute_gap(res['mean_wmax_wmin'], bks['bks_wmax_wmin'])

            ws.append([
                instance,
                res['method'],
                bks['bks_wmax'],
                bks['bks_wmax_wmin'],
                res['best_wmax'],
                res['mean_wmax'],
                gap_wmax,
                res['best_wmax_wmin'],
                res['mean_wmax_wmin'],
                gap_wmax_wmin,
                res['time_sec']
            ])

    wb.save('analysis/comparison_table/comparison_table.xlsx')
    print("✅ Comparison report saved to 'analysis/comparison_table/comparison_table.xlsx'")

if __name__ == '__main__':
    main()
